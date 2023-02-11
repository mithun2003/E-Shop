from django.shortcuts import redirect, render
from .models import *
from cart.models import *
import openpyxl
from datetime import datetime, time, timedelta
import pytz
from account.models import User
from order.models import Order, OrderProduct
from order.forms import *
from django.http import JsonResponse
from .forms import *
#from order.forms import OrderForm
from django.contrib import messages
from django.template import loader
from django.template.defaultfilters import slugify
from .decorators import user_is_admin
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models.functions import ExtractYear, ExtractMonth
from django.db.models import Q
from django.db.models import Count
from django.http import HttpResponse
from django.conf import settings
import calendar
from datetime import datetime
from io import BytesIO
from django.template.loader import get_template
from xhtml2pdf import pisa
import xlwt


# Create your views here.

def dashboard(request):
    
    chart_data = OrderProduct.objects.all()
    print(chart_data)
    salespermonth = [0,0,0,0,0,0,0,0,0,0,0,0]
    for a in range(12):
        for b in chart_data:
            if b.created_at.month == a+1:
                salespermonth[a]+=1
  

    category = Category.objects.all()
    total_order = len(chart_data)
    print(total_order)
    total_purchase_amount = int()
    total_product_qty = int()
    for data in chart_data:
        total_purchase_amount += data.order.order_total
        total_product_qty += data.quantity

    barlist =[]

    for x in category:
        n=0
        for y in chart_data:
            p = Product.objects.get(product_name = y.product.product_name)
            if x.category_name == p.category.category_name:
                print('3#')
                n+=1
        barlist.append(n)
        print(barlist)
        print(salespermonth)

    content = {
        'total_order':total_order,
        'total_purchase_amount':total_purchase_amount,
        'total_product_qty':total_product_qty,
        'barlist':barlist,
        'category':category,
        'salespermonth':salespermonth
    }

    return render(request,'admin/dashboard.html',content)
@user_is_admin
def admin_dashboard(request):
    orders_month = Order.objects.annotate(month=ExtractMonth("created_at")).values("month").annotate(count=Count('id')).values('month','count')
    months=[]
    total_orders = []
    for order in orders_month:
        months.append(calendar.month_name[order['month']])
        total_orders.append(order['count'])
    print(order_detail)
    orders = Order.objects.order_by('-created_at')[:2]
    users = User.objects.filter(is_staff = False, is_admin = False, is_superuser = False).order_by('-date_joined')[:5]
    order_details = OrderProduct.objects.order_by('-created_at')
    context = {
        'total_orders': total_orders,
        'months' : months,
        'orders' : orders,
        'users' : users,
        'order_details':order_details,
    }
    return render(request,"admin/dashboard.html")
@user_is_admin
def variation_list(request):
    p = Paginator(Variation.objects.all(),10)
    page = request.GET.get('page')
    variation_details = p.get_page(page)
    return render(request,'admin/variation.html',{'variation':variation_details})


@user_is_admin
def variation_create(request):
    form = VariationForm()
    if request.method == 'POST':
        form = VariationForm(request.POST)
        if form.is_valid():
            data = form.save()
            messages.success(request, "Variation Created")
            return redirect('variation_list')
        else:
            messages.error(request,"Variation creation failed ")

    context ={
        'variation_form':form
    }
    return render(request,"admin/addvariation.html", context)
@user_is_admin
def variation_delete(request, variation_id):
    variation = Variation.objects.get(id=variation_id)
    if request.method == 'POST':
        variation.delete()
        render(request, 'admin/variation.html')
    return redirect(to='variation_list')
#=====CATEGORY=====
@user_is_admin
def category_list(request):
    categories = Category.objects.all()
    if request.method == 'POST':
        keyword = request.POST.get('keyword')
        categories = categories.filter(Q(category_name__icontains = keyword) )
    paginator = Paginator(categories, 8)
    page = request.GET.get('page')
    paged_categories = paginator.get_page(page)
    content = {
        'categories': paged_categories
    }
    return render(request,"admin/category.html", content)
@user_is_admin
def category_create(request):
    form = CategoryForm()
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            data = form.save()
            messages.success(request, "Category Created")
            return redirect('category_list')
        else:
            messages.error(request,"Category creation failed ")

    context ={
        'category_form':form
    }

    return render(request,"admin/addcategory.html", context)

@user_is_admin
def category_update(request, category_id):
    category = Category.objects.get(id = category_id)
    if request.method == 'POST':
        category_form = CategoryForm(request.POST, request.FILES, instance=category)
        if category_form.is_valid():
            category_form.save()
            messages.success(request, 'Category Updated Successfully')
            return redirect(to='category_list')
    else:
        category_form = CategoryForm(instance=category)
        
    return render(request,"admin/category_update.html",  {'category_form': category_form})
@user_is_admin
def category_delete(request, category_id):
    category = Category.objects.get(id=category_id)
    if request.method == 'POST':
        category.delete()
        render(request, 'admin/category.html')
    return redirect(to='category_list')
#=====END CATEGORY=====

#=====PRODUCT=====
@user_is_admin
def product_list(request):
    print('product')
    products = Product.objects.all().order_by('-updated_at')
    if request.method == 'POST':
        keyword = request.POST.get('keyword')
        products = products.filter(Q(name__icontains = keyword) | Q(category__name__icontains = keyword) | Q(description__icontains = keyword)).order_by('-created')
    paginator = Paginator(products, 7)
    page = request.GET.get('page')
    paged_products = paginator.get_page(page)
    content ={
        'product_details': paged_products
    }
    return render(request,'admin/producttable.html', content)


@user_is_admin
def product_create(request):
    product_form = ProductForm()
    product_image_form = ProductImageForm()
    if request.method == 'POST':
        product_form = ProductForm(request.POST or None, request.FILES or None)
        product_image_form = ProductImageForm(request.POST or None, request.FILES or None)
        images = request.FILES.getlist('image')
        print(product_form.errors)
        if product_form.is_valid():
            name = product_form.cleaned_data['product_name']
            category = product_form.cleaned_data['category']
            slug = product_form.cleaned_data['slug']
            main_image = product_form.cleaned_data['main_image']
            description = product_form.cleaned_data['product_desc']
            variation = product_form.cleaned_data['variation']
            price = product_form.cleaned_data['price']
            stock = product_form.cleaned_data['stock']
            is_available = product_form.cleaned_data['is_available']
            offer = product_form.cleaned_data['offer']
            is_offer=product_form.cleaned_data['is_offer']

            product = Product.objects.create(product_name=name, category=category,slug=slug, product_image=main_image, product_desc=description, price=price, is_available=is_available, stock=stock,is_offer=is_offer,offer=offer)
            product.variation.set(variation)
            for image in images:
                ProductImage.objects.create(product=product, image=image)
            messages.success(request, "Product Created")
            return redirect('product_list')
        else:
            messages.error(request,"Product creation failed ")

    context ={
        'product_form': product_form,
        'product_image_form': product_image_form
    }

    return render(request,"admin/product_create.html", context)

@user_is_admin
def product_update(request, product_id):
    product = Product.objects.get(id = product_id)
    if request.method == 'POST':
        product_form = ProductForm(request.POST, request.FILES, instance=product)
        if product_form.is_valid():
            product_form.save()
            messages.success(request, 'Product Updated Successfully')
            return redirect(to='product_list')
    else:
        product_form = ProductForm(instance=product)
        
    return render(request,"admin/product_update.html",  {'product_form': product_form})

@user_is_admin
def product_delete(request):
    """ product = Product.objects.get(id=product_id)
    if request.method == 'POST':
        product.delete()
        return render(request, 'admin/producttable.html')
    #return redirect(to='product_list')
    #return JsonResponse({'value':True}) """
    
    id = request.GET['product_id']
    delete_product = Product.objects.get(id = id)

    delete_product.delete()
    return JsonResponse({'value':True})


#=====END PRODUCT=====


# user table customize functions start

def user_table(request):
    """
    this is user details list function
    """

    # pagination 
    p = Paginator(User.objects.all(),5)
    page = request.GET.get('page')
    user_details = p.get_page(page)

    return render(request,'admin/usertable.html',{'user_details':user_details})
    


def user_block(request):
    '''
        this function user block users
    '''
    id = request.GET['productid']
    block_user_details = User.objects.get(id = id)
    block_user_details.is_block = 1
    block_user_details.save()
    return JsonResponse({'value':True})



def user_unblock(request,id):
    '''
        this function user unblock the user
    '''

    block_user_details = User.objects.get(id = id)
    block_user_details.is_block = 0
    block_user_details.save()
    return redirect('user_table')


# banner start 
def banners(request):
    banner_details = Banner.objects.all()
    return render(request,'admin/banner.html',{'banner_details':banner_details})

def add_banner(request):
    forms = AddBanner(request.POST or None,request.FILES )
    if request.method == 'POST':
        if forms.is_valid():

            print(forms)
            forms.save()

            return redirect('banners')
    return render(request,'admin/addbanner.html',{'forms':forms})
def banner_delete(request,banner_id):

    banner = Banner.objects.get(id=banner_id)
    banner.delete()
    return render(request, 'admin/banner.html')
    #return redirect(to='product_list')
# banner end
@user_is_admin
def order_list(request):
    date_form = DateForm()
    orders = Order.objects.all().order_by('-created_at')
    if request.method == 'POST':
        keyword = request.POST.get('keyword')
        orders = orders.filter(Q(status__icontains = keyword) | Q(order_number__icontains = keyword) |  Q(user__email__icontains = keyword) | Q(payment__payment_method__icontains = keyword)).order_by('-created_at')
    paginator = Paginator(orders, 15)
    page = request.GET.get('page')
    paged_orders = paginator.get_page(page)
    content ={
        'orders': paged_orders,
        'date_form': date_form
    }
    return render(request,"admin/order.html", content)

def order_history(request):
    '''
    order history
    '''
    p = Paginator(OrderProduct.objects.all().order_by('-created_at'),5)
    page = request.GET.get('page')
    order_details = p.get_page(page)
    return render(request,'admin/orderhistory.html',{'order_details':order_details})

@user_is_admin
def order_update(request, order_number):
    
    order = Order.objects.get(order_number = order_number)
    print(order)
    if request.method == 'POST':
        order_form = OrderStatusForm(request.POST, instance=order)
        if order_form.is_valid():
            order_form.save()
        return redirect('order_list')
    else:
        order_form = OrderStatusForm(instance=order)
    #products=Product.objects.filter(order)
    print(order)
    order_items = OrderProduct.objects.filter(order_id=order.id)
    print(order_items)
    context = {
        'order': order,
        'order_items': order_items,
        'order_form': order_form,
    }
    return render (request, 'admin/order_update.html', context)




def sales_report(request):
    print(request.method)
    if request.method == 'POST':    
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        order = Order.objects.all()
        generate = request.POST.get('generate')
        print(generate) 
        print(start_date)
        
        if generate == 'PDF':
            if start_date and end_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                order = order.order_by('-created_at').filter(created_at__range=[start_date, end_date])
            else:
                order = Order.objects.all().order_by('-created_at')
            
            now=datetime.today()
            data={
                'start_date':start_date,
                'end_date':end_date,
                'orders': order,
            }
            response = HttpResponse(content_type='application/pdf')
            filename = "Report"+str(now)+ ".pdf"
            content = "attachment; filename="+filename
            response['Content-Disposition'] = content
            template = get_template("admin/order_de.html")
            html = template.render(data)
            result = BytesIO()
            pdf = pisa.pisaDocument( BytesIO(html.encode("ISO-8859-1")), result)
            if not pdf.err:
                return HttpResponse(result.getvalue(), content_type='application/pdf')
            return response
        
        elif generate == 'Excel': 
            if start_date and end_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                default_time = time(hour=0, minute=0, second=0)
                # start_date = datetime.combine(start_date, default_time).replace(tzinfo=tzinfo)
                # end_date = datetime.combine(end_date, default_time).replace(tzinfo=tzinfo)
                print(start_date)
                print(end_date)
                orders = order.order_by('-created_at').filter(created_at__range=[start_date, end_date])
            else:
                orders = Order.objects.all().order_by('-created_at')
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['Order Number', 'User' , 'Order Status' , 'Order Total' , 'Order Date',])
            for order in orders:
                date = order.created_at.astimezone(pytz.utc).replace(tzinfo=None)
                #payment = order.payment.payment_method if order.payment.payment_method else "Nil"
                ws.append([order.order_number,order.user.email,order.status, order.order_total,date,])
            file_name = "sales_report.xlsx"
            wb.save(file_name)
            with open(file_name, "rb") as f:
                response = HttpResponse(f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                response["Content-Disposition"] = f"attachment; filename={file_name}"
                return response
            print(response)
            # return response

    else:
        return redirect('order_list')


def export_sales_xls(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        order = Order.objects.all()
        
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            order = order.order_by('-created_at').filter(created_at__range=[start_date, end_date])
        else:
            order = Order.objects.all().order_by('-created_at')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Start Date'
        ws['B1'] = 'End Date'
        ws['A2'] = start_date
        ws['B2'] = end_date
        ws.append(['Order Number', 'Order Date', 'Order Total'])
        for order in orders:
            ws.append([order.order_number, order.created_at, order.total])
        wb.save(response)
        return response
    else:
        return redirect('order_list')



#=====COUPON=====
@user_is_admin
def coupon_list(request):
    coupon = Coupon.objects.all()
    if request.method == 'POST':
        keyword = request.POST.get('keyword')
        coupon = coupon.filter(Q(code__icontains = keyword) )
    paginator = Paginator(coupon, 8)
    page = request.GET.get('page')
    paged_coupon = paginator.get_page(page)
    content = {
        'coupon': paged_coupon
    }
    return render(request,"admin/coupon.html", content)
@user_is_admin
def coupon_create(request):
    form = CouponForm()
    if request.method == 'POST':
        form = CouponForm(request.POST)
        print(form.errors)
        if form.is_valid():
            data = form.save()
            messages.success(request, "Coupon Created")
            return redirect(to='coupon_list')
        else:
            messages.error(request,"Coupon creation failed ")

    context ={
        'coupon_form':form
    }

    return render(request,"admin/addcoupon.html", context)

@user_is_admin
def coupon_update(request, coupon_id):
    coupon = Coupon.objects.get(id = coupon_id)
    if request.method == 'POST':
        coupon_form = CouponForm(request.POST,  instance=coupon)
        if coupon_form.is_valid():
            coupon_form.save()
            messages.success(request, 'Category Updated Successfully')
            return redirect(to='coupon_list')
    else:
        coupon_form = CouponForm(instance=coupon)
        
    return render(request,"admin/couponupdate.html",  {'coupon_form': coupon_form})
@user_is_admin
def coupon_delete(request, coupon_id):
    coupon = Coupon.objects.get(id=coupon_id)
    if request.method == 'POST':
        coupon.delete()
        render(request, 'admin/coupon.html')
    return redirect(to='coupon_list')
#=====END COUPON=====